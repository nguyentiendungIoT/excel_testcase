from __future__ import annotations

from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

TARGET_COLUMN_INDEX = 11  # Column K
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}


def column_width_to_pixels(width: Optional[float]) -> int:
    """Convert Excel column width units to an approximate pixel width."""
    if width is None:
        width = 8.43  # Excel default column width
    if width <= 0:
        return 0
    if width < 1:
        return int(width * 12 + 0.5)
    return int(width * 7 + 5)


def row_height_to_pixels(height: Optional[float]) -> int:
    """Convert Excel row height points to pixels."""
    if height is None:
        height = 15.0  # Excel default row height in points
    if height <= 0:
        return 0
    return int(height * 96 / 72)


def get_image_anchor_row_col(image) -> Tuple[Optional[int], Optional[int]]:
    """Extract top-left anchor cell (row, column) for an image."""
    anchor = image.anchor

    if isinstance(anchor, str):
        row, col = coordinate_to_tuple(anchor)
        return row, col

    if hasattr(anchor, "_from"):
        # openpyxl stores row/col as zero-based in drawing anchors
        return anchor._from.row + 1, anchor._from.col + 1

    return None, None


def fit_images_in_column_k(workbook_path: Path) -> Tuple[int, int]:
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, keep_vba=keep_vba)

    updated_images = 0
    total_images = 0
    col_letter = get_column_letter(TARGET_COLUMN_INDEX)

    for ws in wb.worksheets:
        images = getattr(ws, "_images", [])
        total_images += len(images)

        for image in images:
            row, col = get_image_anchor_row_col(image)
            if row is None or col is None:
                continue
            if col != TARGET_COLUMN_INDEX:
                continue

            cell_width_px = column_width_to_pixels(ws.column_dimensions[col_letter].width)
            cell_height_px = row_height_to_pixels(ws.row_dimensions[row].height)

            if cell_width_px <= 0 or cell_height_px <= 0:
                continue

            target_cell = ws.cell(row=row, column=TARGET_COLUMN_INDEX).coordinate
            image.anchor = target_cell
            image.width = cell_width_px
            image.height = cell_height_px
            updated_images += 1

    wb.save(workbook_path)
    return updated_images, total_images


def main() -> None:
    root = Path(__file__).resolve().parent
    excel_files = [
        p
        for p in root.iterdir()
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS and not p.name.startswith("~$")
    ]

    if not excel_files:
        print("Khong tim thay file Excel nao de xu ly.")
        return

    grand_updated = 0
    grand_total_images = 0

    for excel_file in sorted(excel_files):
        try:
            updated, total_images = fit_images_in_column_k(excel_file)
            grand_updated += updated
            grand_total_images += total_images
            print(f"{excel_file.name}: da can chinh {updated} anh (tong anh trong file: {total_images})")
        except Exception as exc:
            print(f"{excel_file.name}: loi -> {exc}")

    print("---")
    print(f"Tong so anh da can chinh o cot K: {grand_updated}")
    print(f"Tong so anh duoc phat hien trong cac file: {grand_total_images}")


if __name__ == "__main__":
    main()
